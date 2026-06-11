"""Professional Excel export builders for the laundry vertical.

Writes artefacts under ``<repo_root>/exports/laundry/`` and returns metadata
for persistence in ``laundry_exports``.
"""
from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from laundry.normalization import ground_floor_status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EXPORT_FORMATS = ("excel",)
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_TITLE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
_TITLE_FILL = PatternFill("solid", fgColor="0F172A")
_ROW_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_VERDICT_FILLS = {
    "approved_candidate": PatternFill("solid", fgColor="DCFCE7"),
    "manual_review": PatternFill("solid", fgColor="FEF9C3"),
    "rejected": PatternFill("solid", fgColor="FEE2E2"),
    "extraction_failed": PatternFill("solid", fgColor="FFEDD5"),
}

_PIPELINE_COLUMNS: List[Tuple[str, str, Optional[str], float]] = [
    ("Property Name", "property_name", None, 28),
    ("Address", "address", None, 32),
    ("Neighbourhood", "neighbourhood", None, 18),
    ("Acquisition Type", "acquisition_type", None, 14),
    ("Property Type", "property_type", None, 16),
    ("Ground Floor", "ground_floor_status", None, 18),
    ("Area (m²)", "floor_area_m2", "#,##0.0", 12),
    ("Price (€)", "asking_price", "€#,##0", 14),
    ("Rent (€/mo)", "asking_rent_month", "€#,##0", 14),
    ("Revenue (€/yr)", "expected_revenue_eur", "€#,##0", 16),
    ("EBITDA (€)", "ebitda_eur", "€#,##0", 14),
    ("Margin", "operating_margin", "0.0%", 10),
    ("Payback (yr)", "payback_years", "0.0", 12),
    ("Score", "score", "0", 8),
    ("Verdict", "verdict", None, 14),
    ("Status", "deal_status", None, 16),
    ("Risk Count", "risk_count", "0", 10),
    ("Listing URL", "listing_url", None, 42),
    ("Scan Date", "scan_date", None, 20),
]


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def export_dir() -> Path:
    target = repo_root() / "exports" / "laundry"
    target.mkdir(parents=True, exist_ok=True)
    return target


def _slug(text: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9_-]+", "-", text or "laundry")
    return text.strip("-")[:48] or "laundry"


def _stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def _money(v: Any) -> Optional[float]:
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _pct(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        f = float(v)
        return f / 100.0 if f > 1 else f
    except (TypeError, ValueError):
        return None


def _text(v: Any, limit: int = 8000) -> str:
    if v is None:
        return ""
    s = str(v)
    return s if len(s) <= limit else s[: limit - 3] + "..."


def _write_header(ws, headers: List[str], *, row: int = 1) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border = _ROW_BORDER


def _autosize(ws, widths: List[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _flatten_property_row(
    prop: Dict[str, Any],
    analysis: Optional[Dict[str, Any]] = None,
    *,
    job: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    analysis = analysis or {}
    economics = analysis.get("economics") or {}
    scoring = analysis.get("score") or {}
    dd = analysis.get("due_diligence") or {}
    extracted = analysis.get("input") or prop.get("raw_extracted") or {}
    if isinstance(extracted, str):
        try:
            extracted = json.loads(extracted)
        except Exception:
            extracted = {}

    risks = dd.get("red_flags") or dd.get("risks") or []
    if not isinstance(risks, list):
        risks = [str(risks)] if risks else []

    name = (
        prop.get("address")
        or extracted.get("title")
        or prop.get("neighbourhood")
        or         f"Listing {str(prop.get('id', ''))[:8]}"
    )

    gf = ground_floor_status({**extracted, **{k: prop.get(k) for k in ("ground_floor",) if prop.get(k) is not None}})

    return {
        "property_id": prop.get("id"),
        "property_name": name,
        "address": prop.get("address") or extracted.get("address"),
        "city": prop.get("city") or extracted.get("city"),
        "neighbourhood": prop.get("neighbourhood") or extracted.get("neighbourhood"),
        "acquisition_type": prop.get("acquisition_type") or extracted.get("acquisition_type"),
        "property_type": prop.get("property_type") or extracted.get("property_type"),
        "ground_floor_status": gf.get("label"),
        "ground_floor": gf.get("ground_floor"),
        "floor_area_m2": _money(prop.get("floor_area_m2") or economics.get("floor_area_m2") or extracted.get("floor_area_m2")),
        "asking_price": _money(prop.get("asking_price") or extracted.get("asking_price")),
        "asking_rent_month": _money(prop.get("asking_rent_month") or extracted.get("asking_rent_month")),
        "expected_revenue_eur": _money(prop.get("expected_revenue_eur") or economics.get("expected_revenue_eur")),
        "ebitda_eur": _money(prop.get("ebitda_eur") or economics.get("ebitda_eur")),
        "operating_margin": _pct(prop.get("operating_margin") or economics.get("operating_margin")),
        "payback_years": _money(prop.get("payback_years") or economics.get("payback_years")),
        "score": prop.get("score") if prop.get("score") is not None else scoring.get("score"),
        "verdict": prop.get("verdict") or scoring.get("verdict"),
        "deal_status": prop.get("deal_status") or scoring.get("deal_status"),
        "classification": prop.get("classification") or scoring.get("classification"),
        "risk_count": len(risks),
        "listing_url": prop.get("listing_url"),
        "scan_date": (job or {}).get("finished_at") or (job or {}).get("created_at") or prop.get("created_at"),
        "search_url": (job or {}).get("search_url"),
        "job_id": prop.get("job_id") or (job or {}).get("id"),
        "memo_md": analysis.get("ic_memo") or "",
        "extracted": extracted,
        "economics": economics,
        "due_diligence": dd,
        "location": analysis.get("location") or {},
        "risks": risks,
    }


def build_pipeline_workbook(rows: List[Dict[str, Any]], *, title: str = "Pipeline") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"
    ws.freeze_panes = "A2"

    headers = [c[0] for c in _PIPELINE_COLUMNS]
    accessors = [c[1] for c in _PIPELINE_COLUMNS]
    formats = [c[2] for c in _PIPELINE_COLUMNS]
    widths = [c[3] for c in _PIPELINE_COLUMNS]

    _write_header(ws, headers)
    for row_idx, record in enumerate(rows, start=2):
        for col_idx, (key, fmt) in enumerate(zip(accessors, formats), start=1):
            value = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else None)
            if fmt and isinstance(value, (int, float)):
                cell.number_format = fmt
            cell.alignment = _ALIGN_LEFT if not fmt else _ALIGN_CENTER
            cell.border = _ROW_BORDER
        status_cell = ws.cell(row=row_idx, column=16)
        fill = _VERDICT_FILLS.get(record.get("deal_status") or "")
        if fill:
            status_cell.fill = fill

    if rows:
        end_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{end_col}{len(rows) + 1}"
    _autosize(ws, widths)

    summary = wb.create_sheet("Summary")
    summary["A1"] = f"Laundry Pipeline Export — {title}"
    summary["A1"].font = _TITLE_FONT
    summary["A1"].fill = _TITLE_FILL
    summary.merge_cells("A1:D1")
    summary.append([])
    summary.append(["Generated", datetime.now(timezone.utc).isoformat()])
    summary.append(["Properties", len(rows)])
    summary.append(["Approved", len([r for r in rows if r.get("deal_status") == "approved_candidate"])])
    summary.append(["Manual Review", len([r for r in rows if r.get("deal_status") == "manual_review"])])
    summary.append(["Rejected", len([r for r in rows if r.get("deal_status") == "rejected"])])
    summary.append(["Extraction Failed", len([r for r in rows if r.get("deal_status") == "extraction_failed"])])
    _autosize(summary, [22, 18, 18, 18])
    return wb


def _kv_sheet(ws, title: str, pairs: List[Tuple[str, Any]]) -> None:
    ws.title = title[:31]
    ws.freeze_panes = "A2"
    _write_header(ws, ["Field", "Value"])
    for row_idx, (label, value) in enumerate(pairs, start=2):
        ws.cell(row=row_idx, column=1, value=label).border = _ROW_BORDER
        cell = ws.cell(row=row_idx, column=2, value=_text(value) if not isinstance(value, (int, float)) else value)
        cell.alignment = _ALIGN_LEFT
        cell.border = _ROW_BORDER
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 72


def build_single_deal_workbook(
    prop: Dict[str, Any],
    analysis: Optional[Dict[str, Any]] = None,
    *,
    job: Optional[Dict[str, Any]] = None,
) -> Workbook:
    row = _flatten_property_row(prop, analysis, job=job)
    economics = row.get("economics") or {}
    dd = row.get("due_diligence") or {}
    extracted = row.get("extracted") or {}

    wb = Workbook()

    exec_ws = wb.active
    exec_ws.title = "Executive Summary"
    exec_ws["A1"] = "Executive Summary"
    exec_ws["A1"].font = _TITLE_FONT
    exec_ws["A1"].fill = _TITLE_FILL
    exec_ws.merge_cells("A1:B1")
    exec_pairs = [
        ("Property", row.get("property_name")),
        ("Address", row.get("address")),
        ("Neighbourhood", row.get("neighbourhood")),
        ("City", row.get("city")),
        ("Acquisition", row.get("acquisition_type")),
        ("Property Type", row.get("property_type")),
        ("Ground Floor", row.get("ground_floor_status")),
        ("Floor Area (m²)", row.get("floor_area_m2")),
        ("Score", row.get("score")),
        ("Verdict", row.get("verdict")),
        ("Classification", row.get("classification")),
        ("Deal Status", row.get("deal_status")),
        ("Investment Recommendation", row.get("verdict") or "See IC memo"),
    ]
    _write_header(exec_ws, ["Section", "Detail"], row=3)
    for idx, (k, v) in enumerate(exec_pairs, start=4):
        exec_ws.cell(row=idx, column=1, value=k).border = _ROW_BORDER
        cell = exec_ws.cell(row=idx, column=2, value=v)
        cell.border = _ROW_BORDER
        if k == "Deal Status":
            fill = _VERDICT_FILLS.get(str(v or ""))
            if fill:
                cell.fill = fill
    exec_ws.column_dimensions["A"].width = 28
    exec_ws.column_dimensions["B"].width = 56

    fin_pairs = [
        ("Expected Revenue (€/yr)", _money(economics.get("expected_revenue_eur") or row.get("expected_revenue_eur"))),
        ("Utilities (€/yr)", _money(economics.get("utilities_eur") or economics.get("annual_utilities"))),
        ("Rent (€/yr)", _money(economics.get("annual_rent") or (row.get("asking_rent_month") or 0) * 12 if row.get("asking_rent_month") else None)),
        ("Payroll (€/yr)", _money(economics.get("payroll_eur") or economics.get("annual_payroll"))),
        ("Maintenance (€/yr)", _money(economics.get("maintenance_eur") or economics.get("annual_maintenance"))),
        ("EBITDA (€)", _money(economics.get("ebitda_eur") or row.get("ebitda_eur"))),
        ("Operating Margin", _pct(economics.get("operating_margin") or row.get("operating_margin"))),
        ("Cashflow (€/yr)", _money(economics.get("cashflow_eur") or economics.get("ebitda_eur"))),
        ("Payback (years)", _money(economics.get("payback_years") or row.get("payback_years"))),
        ("ROI", _pct(economics.get("roi") or economics.get("ebitda_yield"))),
        ("Asking Price (€)", row.get("asking_price")),
        ("Asking Rent (€/mo)", row.get("asking_rent_month")),
    ]
    fin_ws = wb.create_sheet("Financial Model")
    _write_header(fin_ws, ["Metric", "Value"])
    for idx, (k, v) in enumerate(fin_pairs, start=2):
        fin_ws.cell(row=idx, column=1, value=k).border = _ROW_BORDER
        cell = fin_ws.cell(row=idx, column=2, value=v)
        cell.border = _ROW_BORDER
        if k.endswith("(€)") or "€" in k:
            if isinstance(v, (int, float)):
                cell.number_format = "€#,##0"
        if k in ("Operating Margin", "ROI"):
            if isinstance(v, (int, float)):
                cell.number_format = "0.0%"
    fin_ws.column_dimensions["A"].width = 28
    fin_ws.column_dimensions["B"].width = 20

    risk_ws = wb.create_sheet("Risk Assessment")
    risk_ws.freeze_panes = "A2"
    _write_header(risk_ws, ["Category", "Detail"])
    r = 2
    for item in row.get("risks") or []:
        risk_ws.cell(row=r, column=1, value="Risk").border = _ROW_BORDER
        risk_ws.cell(row=r, column=2, value=_text(item)).border = _ROW_BORDER
        r += 1
    for section, items in (dd or {}).items():
        if section in ("red_flags", "risks"):
            continue
        if isinstance(items, list):
            for item in items:
                risk_ws.cell(row=r, column=1, value=section).border = _ROW_BORDER
                risk_ws.cell(row=r, column=2, value=_text(item)).border = _ROW_BORDER
                r += 1
        elif items:
            risk_ws.cell(row=r, column=1, value=section).border = _ROW_BORDER
            risk_ws.cell(row=r, column=2, value=_text(items)).border = _ROW_BORDER
            r += 1
    checklist = dd.get("checklist") or dd.get("due_diligence_checklist") or []
    if isinstance(checklist, list):
        for item in checklist:
            risk_ws.cell(row=r, column=1, value="Checklist").border = _ROW_BORDER
            risk_ws.cell(row=r, column=2, value=_text(item)).border = _ROW_BORDER
            r += 1
    risk_ws.column_dimensions["A"].width = 22
    risk_ws.column_dimensions["B"].width = 72

    prop_pairs: List[Tuple[str, Any]] = []
    for k, v in {**extracted, **{k: prop.get(k) for k in prop if k != "raw_extracted"}}.items():
        if isinstance(v, (dict, list)):
            prop_pairs.append((k, json.dumps(v, default=str)))
        else:
            prop_pairs.append((k, v))
    prop_ws = wb.create_sheet("Property Data")
    _kv_sheet(prop_ws, "Property Data", prop_pairs)

    source_pairs = [
        ("Listing URL", row.get("listing_url")),
        ("Search URL", row.get("search_url")),
        ("Scan Job ID", row.get("job_id")),
        ("Scan Date", row.get("scan_date")),
        ("Property ID", row.get("property_id")),
        ("Export Generated", datetime.now(timezone.utc).isoformat()),
    ]
    source_ws = wb.create_sheet("Source Information")
    _kv_sheet(source_ws, "Source Information", source_pairs)

    return wb


def save_workbook(wb: Workbook, filename: str) -> Dict[str, Any]:
    path = export_dir() / filename
    wb.save(path)
    return {
        "format": "excel",
        "file_path": str(path),
        "filename": path.name,
        "size_bytes": path.stat().st_size,
        "mime_type": MIME_XLSX,
    }


def filename_for_single(prop: Dict[str, Any]) -> str:
    base = _slug(prop.get("address") or str(prop.get("id", "deal"))[:8])
    return f"laundry_deal_{base}_{_stamp()}.xlsx"


def filename_for_pipeline(label: str, count: int) -> str:
    return f"laundry_pipeline_{_slug(label)}_{count}_{_stamp()}.xlsx"


def build_pipeline_export_rows(
    properties: List[Dict[str, Any]],
    analyses_by_prop: Optional[Dict[str, Dict[str, Any]]] = None,
    *,
    job: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    analyses_by_prop = analyses_by_prop or {}
    rows: List[Dict[str, Any]] = []
    for prop in properties:
        pid = prop.get("id")
        analysis = analyses_by_prop.get(pid) or prop.get("latest_analysis")
        rows.append(_flatten_property_row(prop, analysis, job=job))
    return rows


def generate_single_deal_export(
    prop: Dict[str, Any],
    analysis: Optional[Dict[str, Any]] = None,
    *,
    job: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    wb = build_single_deal_workbook(prop, analysis, job=job)
    return save_workbook(wb, filename_for_single(prop))


def generate_pipeline_export(
    properties: List[Dict[str, Any]],
    analyses_by_prop: Optional[Dict[str, Dict[str, Any]]] = None,
    *,
    label: str = "Pipeline",
    job: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    rows = build_pipeline_export_rows(properties, analyses_by_prop, job=job)
    wb = build_pipeline_workbook(rows, title=label)
    meta = save_workbook(wb, filename_for_pipeline(label, len(rows)))
    meta["row_count"] = len(rows)
    meta["label"] = label
    return meta


_LISTING_ACCOUNTING_COLUMNS: List[Tuple[str, str, Optional[str], float]] = [
    ("Listing URL", "listing_url", None, 42),
    ("Index", "listing_index", "0", 8),
    ("Status", "status", None, 18),
    ("Reason Code", "reason_code", None, 22),
    ("Reason Message", "reason_message", None, 36),
    ("Stage Failed", "stage_failed", None, 16),
    ("Attempt Count", "attempt_count", "0", 12),
    ("Property ID", "property_id", None, 38),
    ("Duplicate Of", "duplicate_of_property_id", None, 38),
    ("Filter Name", "filter_name", None, 18),
    ("Filter Value", "filter_value", None, 14),
    ("Actual Value", "actual_value", None, 14),
    ("Score", "score", "0", 8),
    ("Verdict", "verdict", None, 16),
    ("Deal Status", "deal_status", None, 18),
    ("Address", "address", None, 28),
    ("Title", "title", None, 24),
]


def _flatten_listing_result_row(row: Dict[str, Any]) -> Dict[str, Any]:
    result = row.get("result") or {}
    if isinstance(result, str):
        try:
            result = json.loads(result)
        except Exception:
            result = {}
    return {
        "listing_url": row.get("listing_url") or result.get("listing_url"),
        "listing_index": row.get("listing_index"),
        "status": row.get("status") or result.get("status") or result.get("terminal_status"),
        "reason_code": row.get("reason_code") or result.get("reason_code"),
        "reason_message": row.get("reason_message") or result.get("reason_message") or row.get("error_message"),
        "stage_failed": row.get("stage_failed") or result.get("stage_failed"),
        "attempt_count": row.get("attempt_count") or result.get("attempt_count") or 0,
        "property_id": row.get("property_id") or result.get("property_id"),
        "duplicate_of_property_id": row.get("duplicate_of_property_id") or result.get("duplicate_of_property_id"),
        "filter_name": row.get("filter_name") or result.get("filter_name"),
        "filter_value": row.get("filter_value") or result.get("filter_value"),
        "actual_value": row.get("actual_value") or result.get("actual_value"),
        "score": row.get("score") or result.get("score"),
        "verdict": result.get("verdict"),
        "deal_status": row.get("deal_status") or result.get("deal_status"),
        "address": row.get("address") or result.get("address"),
        "title": row.get("title") or result.get("title"),
    }


def build_listing_accounting_workbook(
    listing_rows: List[Dict[str, Any]],
    *,
    title: str = "All Listings",
    summary: Optional[Dict[str, Any]] = None,
) -> Workbook:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_rows = [
        ["Metric", "Value"],
        ["Requested limit", (summary or {}).get("requested_limit")],
        ["Effective limit", (summary or {}).get("effective_limit")],
        ["Source found", (summary or {}).get("source_found_count") or (summary or {}).get("source_available_count")],
        ["Discovered", (summary or {}).get("discovered_count")],
        ["Queued", (summary or {}).get("queued_count")],
        ["Processed", (summary or {}).get("processed_count")],
        ["Successful", (summary or {}).get("success_count")],
        ["Duplicates", (summary or {}).get("duplicate_count")],
        ["Filtered out", (summary or {}).get("filtered_out_count")],
        ["Failed", (summary or {}).get("failed_count")],
        ["Exported properties", (summary or {}).get("exported_count")],
    ]
    for r_idx, (label, val) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=r_idx, column=1, value=label)
        ws_summary.cell(row=r_idx, column=2, value=val)

    sheets = {
        "All Listings": listing_rows,
        "Successful": [r for r in listing_rows if (r.get("status") or "").lower() == "success"],
        "Manual Review": [r for r in listing_rows if (r.get("deal_status") or "").lower() == "manual_review"],
        "Approved": [r for r in listing_rows if (r.get("deal_status") or "").lower() == "approved_candidate"],
        "Rejected": [r for r in listing_rows if (r.get("deal_status") or "").lower() == "rejected"],
        "Failed": [r for r in listing_rows if (r.get("status") or "").lower() in (
            "failed", "scrape_failed", "extraction_failed", "persistence_failed", "scoring_failed",
        )],
        "Duplicates": [r for r in listing_rows if (r.get("status") or "").lower() == "duplicate"],
        "Filtered Out": [r for r in listing_rows if (r.get("status") or "").lower() == "filtered_out"],
    }

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        headers = [c[0] for c in _LISTING_ACCOUNTING_COLUMNS]
        _write_header(ws, headers)
        flat_rows = [_flatten_listing_result_row(r) for r in rows]
        for r_idx, flat in enumerate(flat_rows, start=2):
            for c_idx, (_, key, fmt, _) in enumerate(_LISTING_ACCOUNTING_COLUMNS, start=1):
                val = flat.get(key)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = _ROW_BORDER
                cell.alignment = _ALIGN_LEFT
                if fmt and val is not None:
                    cell.number_format = fmt
        _autosize(ws, [c[3] for c in _LISTING_ACCOUNTING_COLUMNS])
    return wb


def generate_listing_accounting_export(
    listing_rows: List[Dict[str, Any]],
    *,
    label: str = "Scan Listings",
    summary: Optional[Dict[str, Any]] = None,
    job: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    flat = [_flatten_listing_result_row(r) for r in listing_rows]
    wb = build_listing_accounting_workbook(listing_rows, title=label, summary=summary)
    meta = save_workbook(wb, filename_for_pipeline(label, len(flat)))
    meta["row_count"] = len(flat)
    meta["label"] = label
    meta["job_id"] = (job or {}).get("id")
    return meta
