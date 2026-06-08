"""
Export builders for the laundry vertical.

All formats accept a single property + its latest analysis and write to disk
under ``<repo_root>/exports/laundry/``. The function returns the absolute path
so the API layer can stream the file back to the browser.

Supported formats:

* ``excel`` — full multi-tab workbook (summary, economics, location, SWOT, memo)
* ``csv``   — flat one-row summary
* ``json``  — verbatim DB record (extracted + economics + score + DD + memo)
* ``zip``   — Excel + JSON + memo.md packaged
* ``memo``  — Markdown memo
* ``financial_model`` — economics-only Excel
* ``full_package``    — every artefact + the assumption snapshot
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


EXPORT_FORMATS = (
    "excel",
    "csv",
    "json",
    "zip",
    "memo",
    "financial_model",
    "full_package",
)


def _repo_root() -> Path:
    """Repo root = three levels up from this file (matches storage exporter)."""
    return Path(__file__).resolve().parents[4]


def export_dir() -> Path:
    return _repo_root() / "exports" / "laundry"


def ensure_export_dir() -> Path:
    target = export_dir()
    target.mkdir(parents=True, exist_ok=True)
    return target


def _slug(text: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9_-]+", "-", text or "untitled")
    return text.strip("-")[:60] or "laundry"


def _filename(prop: Dict[str, Any], fmt: str, suffix: str) -> str:
    base = _slug(f"{prop.get('id', 'x')[:8]}_{prop.get('address') or 'laundry'}")
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    return f"laundry_{base}_{stamp}.{suffix}"


# ---------------------------------------------------------------------------
# Excel builders
# ---------------------------------------------------------------------------


_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _flatten(d: Dict[str, Any], prefix: str = "") -> Dict[str, Any]:
    flat: Dict[str, Any] = {}
    for k, v in (d or {}).items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            flat.update(_flatten(v, key))
        else:
            flat[key] = v
    return flat


def _build_full_workbook(prop: Dict[str, Any], analysis: Dict[str, Any]) -> Workbook:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    _style_header(summary, 1, 2)
    for k, v in prop.items():
        summary.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v, default=str)])
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 80

    eco_ws = wb.create_sheet("Economics")
    eco_ws.append(["Metric", "Value"])
    _style_header(eco_ws, 1, 2)
    for k, v in (analysis.get("economics") or {}).items():
        eco_ws.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v, default=str)])
    eco_ws.column_dimensions["A"].width = 36
    eco_ws.column_dimensions["B"].width = 24

    loc_ws = wb.create_sheet("Location")
    loc_ws.append(["Signal", "Value"])
    _style_header(loc_ws, 1, 2)
    for k, v in (analysis.get("location") or {}).items():
        loc_ws.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v, default=str)])
    loc_ws.column_dimensions["A"].width = 36
    loc_ws.column_dimensions["B"].width = 32

    score_ws = wb.create_sheet("Score")
    score_ws.append(["Field", "Value"])
    _style_header(score_ws, 1, 2)
    for k, v in _flatten(analysis.get("score") or {}).items():
        score_ws.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v, default=str)])
    score_ws.column_dimensions["A"].width = 44
    score_ws.column_dimensions["B"].width = 24

    dd_ws = wb.create_sheet("Due Diligence")
    dd_ws.append(["Section", "Item"])
    _style_header(dd_ws, 1, 2)
    dd = analysis.get("due_diligence") or {}
    for section, items in dd.items():
        if isinstance(items, list):
            for item in items:
                dd_ws.append([section, str(item)])
        elif isinstance(items, dict):
            dd_ws.append([section, json.dumps(items, default=str)])
        else:
            dd_ws.append([section, str(items)])
    dd_ws.column_dimensions["A"].width = 26
    dd_ws.column_dimensions["B"].width = 80

    memo_ws = wb.create_sheet("IC Memo")
    memo_ws.column_dimensions["A"].width = 120
    for line in (analysis.get("ic_memo") or "").splitlines() or [""]:
        memo_ws.append([line])

    return wb


def _build_financial_model_workbook(analysis: Dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Model"
    economics = analysis.get("economics") or {}
    ws.append(["Metric", "Value"])
    _style_header(ws, 1, 2)
    for k, v in economics.items():
        ws.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v, default=str)])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 28

    inputs = wb.create_sheet("Assumptions")
    inputs.append(["Path", "Value"])
    _style_header(inputs, 1, 2)
    for k, v in _flatten(analysis.get("assumptions_used") or {}).items():
        inputs.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v, default=str)])
    inputs.column_dimensions["A"].width = 50
    inputs.column_dimensions["B"].width = 22

    return wb


# ---------------------------------------------------------------------------
# CSV / JSON / memo / zip / full package
# ---------------------------------------------------------------------------


def _csv_bytes(prop: Dict[str, Any], analysis: Dict[str, Any]) -> bytes:
    flat = {
        **{f"property.{k}": v for k, v in prop.items() if not isinstance(v, (dict, list))},
        **{f"economics.{k}": v for k, v in (analysis.get("economics") or {}).items() if not isinstance(v, (dict, list))},
        **{f"score.{k}": v for k, v in _flatten(analysis.get("score") or {}).items() if not isinstance(v, (dict, list))},
        **{f"location.{k}": v for k, v in (analysis.get("location") or {}).items() if not isinstance(v, (dict, list))},
    }
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(list(flat.keys()))
    writer.writerow(list(flat.values()))
    return buf.getvalue().encode("utf-8")


def build_export(
    *,
    fmt: str,
    prop: Dict[str, Any],
    analysis: Dict[str, Any],
) -> Dict[str, Any]:
    """Build the export artefact, write it to disk, and return its metadata."""
    fmt = (fmt or "").lower()
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"Unsupported export format: {fmt}")
    target = ensure_export_dir()

    if fmt == "excel":
        wb = _build_full_workbook(prop, analysis)
        path = target / _filename(prop, fmt, "xlsx")
        wb.save(path)
    elif fmt == "financial_model":
        wb = _build_financial_model_workbook(analysis)
        path = target / _filename(prop, fmt, "xlsx")
        wb.save(path)
    elif fmt == "csv":
        path = target / _filename(prop, fmt, "csv")
        path.write_bytes(_csv_bytes(prop, analysis))
    elif fmt == "json":
        path = target / _filename(prop, fmt, "json")
        payload = {"property": prop, "analysis": analysis}
        path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    elif fmt == "memo":
        path = target / _filename(prop, fmt, "md")
        path.write_text(analysis.get("ic_memo") or "(no memo)", encoding="utf-8")
    elif fmt == "zip":
        path = target / _filename(prop, fmt, "zip")
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            wb = _build_full_workbook(prop, analysis)
            xbuf = io.BytesIO()
            wb.save(xbuf)
            zf.writestr("workbook.xlsx", xbuf.getvalue())
            zf.writestr("memo.md", analysis.get("ic_memo") or "(no memo)")
            zf.writestr("payload.json", json.dumps({"property": prop, "analysis": analysis}, indent=2, default=str))
    elif fmt == "full_package":
        path = target / _filename(prop, fmt, "zip")
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            wb = _build_full_workbook(prop, analysis)
            xbuf = io.BytesIO()
            wb.save(xbuf)
            zf.writestr("workbook.xlsx", xbuf.getvalue())

            fm_wb = _build_financial_model_workbook(analysis)
            fmbuf = io.BytesIO()
            fm_wb.save(fmbuf)
            zf.writestr("financial_model.xlsx", fmbuf.getvalue())

            zf.writestr("memo.md", analysis.get("ic_memo") or "(no memo)")
            zf.writestr("payload.json", json.dumps({"property": prop, "analysis": analysis}, indent=2, default=str))
            zf.writestr(
                "summary.csv",
                _csv_bytes(prop, analysis).decode("utf-8"),
            )
            zf.writestr(
                "assumptions.json",
                json.dumps(analysis.get("assumptions_used") or {}, indent=2, default=str),
            )

    return {
        "format": fmt,
        "file_path": str(path),
        "filename": path.name,
        "size_bytes": path.stat().st_size,
    }
