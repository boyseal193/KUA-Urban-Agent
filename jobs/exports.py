"""K.U.A. scan export builders.

Pure functions that turn a job + listing-result set into a downloadable
artifact (bytes). No I/O to disk, no Supabase calls — that is the storage
layer's job. These builders are safe to call from request handlers,
background workers, or one-off scripts.

Outputs:
  * build_excel_bytes  — 4-sheet professional workbook (xlsx)
  * build_csv_bytes    — flat property pipeline (csv)
  * build_json_bytes   — raw structured output (json)
  * build_memo_bytes   — concatenated IC memos (markdown)
  * build_zip_bytes    — all of the above, plus logs/errors, in one zip
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------------------------------------------------------
# Public type aliases
# ---------------------------------------------------------------------------
ExportFormat = str  # "excel" | "csv" | "json" | "memo" | "zip"

MIME_BY_FORMAT: Dict[str, str] = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "json": "application/json; charset=utf-8",
    "memo": "text/markdown; charset=utf-8",
    "zip": "application/zip",
}

EXTENSION_BY_FORMAT: Dict[str, str] = {
    "excel": "xlsx",
    "csv": "csv",
    "json": "json",
    "memo": "md",
    "zip": "zip",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe(value: Any, default: Any = None) -> Any:
    return value if value is not None else default


def _money(value: Any) -> Optional[float]:
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _pct(value: Any) -> Optional[float]:
    """Return value as a 0-1 fraction (Excel will multiply by 100 via format)."""
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _clean_text(value: Any, limit: int = 32_000) -> str:
    if value is None:
        return ""
    s = str(value)
    return s if len(s) <= limit else s[: limit - 3] + "..."


def _safe_filename_segment(value: Any) -> str:
    if not value:
        return "scan"
    s = str(value)
    out = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        elif ch in (" ", "."):
            out.append("-")
    cleaned = "".join(out).strip("-_")
    return cleaned or "scan"


def filename_for(job: Dict[str, Any], fmt: ExportFormat) -> str:
    """Return a stable, human-readable filename for an export."""
    ext = EXTENSION_BY_FORMAT.get(fmt, "bin")
    job_id = _safe_filename_segment((job or {}).get("id"))[:12]
    ts_raw = (job or {}).get("created_at") or (job or {}).get("started_at")
    try:
        if ts_raw:
            ts = datetime.fromisoformat(str(ts_raw).replace("Z", "+00:00"))
        else:
            ts = datetime.now(timezone.utc)
    except Exception:
        ts = datetime.now(timezone.utc)
    return f"kua-scan-{ts.strftime('%Y%m%d-%H%M%S')}-{job_id}.{ext}"


# ---------------------------------------------------------------------------
# Flatten a listing result row → one canonical record for tables.
# ---------------------------------------------------------------------------
def flatten_result(listing_row: Dict[str, Any]) -> Dict[str, Any]:
    """Take one row from scan_listing_results and return a flat record.

    Tolerates partial / missing nested structures (extracted, economics, …).
    Every key is always present so CSV / Excel never errors on a missing
    column for some rows.
    """
    if not isinstance(listing_row, dict):
        listing_row = {}

    result = listing_row.get("result") if isinstance(listing_row.get("result"), dict) else {}
    extracted = result.get("extracted") if isinstance(result.get("extracted"), dict) else {}
    economics = result.get("economics") if isinstance(result.get("economics"), dict) else {}
    score = result.get("score") if isinstance(result.get("score"), dict) else {}
    auto_scores = result.get("auto_scores") if isinstance(result.get("auto_scores"), dict) else {}

    return {
        "listing_index": listing_row.get("listing_index"),
        "status": listing_row.get("status") or ("ok" if result.get("success") else "failed"),
        "property_id": result.get("property_id"),
        "source_url": listing_row.get("listing_url")
        or result.get("source_url")
        or extracted.get("listing_url"),
        "deal_status": result.get("deal_status"),
        "score": score.get("score"),
        "verdict": score.get("verdict"),
        "classification": score.get("classification"),
        "deal_killer": score.get("deal_killer"),
        # --- Address / location ---
        "address": extracted.get("address"),
        "district": extracted.get("neighbourhood") or extracted.get("district"),
        "city": extracted.get("city") or "Barcelona",
        "latitude": extracted.get("latitude"),
        "longitude": extracted.get("longitude"),
        # --- Size ---
        "gba_m2": _money(extracted.get("gba_m2")),
        "nra_m2": _money(economics.get("nra_m2")),
        "nra_efficiency": _pct(economics.get("nra_efficiency")),
        # --- Pricing ---
        "asking_price": _money(extracted.get("asking_price")),
        "asking_rent_month": _money(extracted.get("asking_rent_month")),
        "rent_per_m2": _money(extracted.get("rent_per_m2")),
        "price_per_m2_nra": _money(extracted.get("price_per_m2_nra")),
        # --- Building ---
        "ceiling_height": extracted.get("ceiling_height"),
        "loading_access": extracted.get("loading_access"),
        "access_type": extracted.get("access_type"),
        "floor_level": extracted.get("floor_level"),
        "building_type": extracted.get("building_type"),
        "current_use": extracted.get("current_use"),
        # --- Economics ---
        "model_type": economics.get("model_type"),
        "estimated_units": economics.get("estimated_units"),
        "monthly_revenue": _money(economics.get("monthly_revenue")),
        "annual_revenue": _money(economics.get("annual_revenue")),
        "annual_rent": _money(economics.get("annual_rent")),
        "annual_opex": _money(economics.get("annual_opex")),
        "ebitda": _money(economics.get("ebitda")),
        "margin": _pct(economics.get("margin")),
        "ebitda_yield": _pct(economics.get("ebitda_yield")),
        "true_ebitda_yield": _pct(economics.get("true_ebitda_yield")),
        "payback_years": _money(economics.get("payback_years")),
        "true_payback_years": _money(economics.get("true_payback_years")),
        "conversion_capex": _money(economics.get("conversion_capex")),
        "total_investment": _money(economics.get("total_investment")),
        # --- Sub-scores ---
        "location_score": auto_scores.get("location_score"),
        "building_score": auto_scores.get("building_score"),
        "economics_score": auto_scores.get("economics_score"),
        "risk_score": auto_scores.get("risk_score"),
        "strategic_fit_score": auto_scores.get("strategic_fit_score"),
        # --- Notes / diagnostics ---
        "due_diligence_flags": " | ".join(score.get("due_diligence_flags") or []) if isinstance(score.get("due_diligence_flags"), list) else "",
        "description": _clean_text(extracted.get("description"), limit=2000),
        "error": listing_row.get("error_message") or result.get("error"),
        "ic_memo": _clean_text(result.get("ic_memo"), limit=20_000),
        "created_at": listing_row.get("created_at"),
        "updated_at": listing_row.get("updated_at"),
    }


# ---------------------------------------------------------------------------
# Aggregated metrics used by the Executive Summary sheet.
# ---------------------------------------------------------------------------
def _aggregate_metrics(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    approved = [r for r in rows if r.get("deal_status") == "approved_candidate"]
    manual = [r for r in rows if r.get("deal_status") == "manual_review"]
    rejected = [r for r in rows if r.get("deal_status") == "rejected"]
    failed = [r for r in rows if r.get("status") == "failed" or (r.get("status") not in {"ok", "running"} and not r.get("property_id"))]

    def _sum(field: str, source: Iterable[Dict[str, Any]]) -> Optional[float]:
        values = [r.get(field) for r in source if isinstance(r.get(field), (int, float))]
        if not values:
            return None
        return float(sum(values))

    def _avg(field: str, source: Iterable[Dict[str, Any]]) -> Optional[float]:
        values = [r.get(field) for r in source if isinstance(r.get(field), (int, float))]
        if not values:
            return None
        return float(sum(values)) / len(values)

    interesting = approved + manual
    return {
        "approved_count": len(approved),
        "manual_review_count": len(manual),
        "rejected_count": len(rejected),
        "failed_count": len(failed),
        "total_count": len(rows),
        "top_opportunities": sorted(
            interesting,
            key=lambda r: r.get("score") or 0,
            reverse=True,
        )[:10],
        "total_ebitda": _sum("ebitda", interesting),
        "total_investment": _sum("total_investment", interesting),
        "avg_margin": _avg("margin", interesting),
        "avg_ebitda_yield": _avg("ebitda_yield", interesting),
        "avg_true_ebitda_yield": _avg("true_ebitda_yield", interesting),
        "avg_payback_years": _avg("true_payback_years", interesting),
    }


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1f2937")  # gunmetal
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Inter", size=11)
_TITLE_FONT = Font(bold=True, color="FFFFFF", name="Inter", size=14)
_TITLE_FILL = PatternFill("solid", fgColor="0f172a")
_ROW_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# Color coding for the deal_status column
_DEAL_STATUS_FILLS = {
    "approved_candidate": PatternFill("solid", fgColor="DCFCE7"),  # green-100
    "manual_review": PatternFill("solid", fgColor="FEF9C3"),       # yellow-100
    "rejected": PatternFill("solid", fgColor="FEE2E2"),            # red-100
}


# (header, accessor, number_format, width)
_PIPELINE_COLUMNS: List[Tuple[str, str, Optional[str], int]] = [
    ("#", "listing_index", "0", 5),
    ("Status", "status", None, 12),
    ("Deal Status", "deal_status", None, 18),
    ("Score", "score", "0", 7),
    ("Verdict", "verdict", None, 16),
    ("Classification", "classification", None, 22),
    ("Deal Killer", "deal_killer", None, 18),
    ("Address", "address", None, 32),
    ("District", "district", None, 18),
    ("City", "city", None, 14),
    ("GBA m²", "gba_m2", "#,##0", 10),
    ("NRA m²", "nra_m2", "#,##0", 10),
    ("NRA Eff.", "nra_efficiency", "0.0%", 9),
    ("Asking Price", "asking_price", '€#,##0', 16),
    ("Ask Rent / mo", "asking_rent_month", '€#,##0', 14),
    ("€/m² NRA", "price_per_m2_nra", '€#,##0', 11),
    ("Ceiling (m)", "ceiling_height", "0.00", 11),
    ("Loading", "loading_access", None, 10),
    ("Access", "access_type", None, 14),
    ("Floor", "floor_level", None, 10),
    ("Building", "building_type", None, 18),
    ("Use", "current_use", None, 16),
    ("Units", "estimated_units", "0", 8),
    ("Monthly Rev.", "monthly_revenue", '€#,##0', 14),
    ("Annual Rev.", "annual_revenue", '€#,##0', 14),
    ("Annual OpEx", "annual_opex", '€#,##0', 14),
    ("EBITDA", "ebitda", '€#,##0', 14),
    ("Margin", "margin", "0.0%", 9),
    ("EBITDA Yield", "ebitda_yield", "0.00%", 12),
    ("True Yield", "true_ebitda_yield", "0.00%", 11),
    ("Payback (yr)", "true_payback_years", "0.0", 11),
    ("CapEx", "conversion_capex", '€#,##0', 14),
    ("Total Inv.", "total_investment", '€#,##0', 14),
    ("Loc Score", "location_score", "0", 9),
    ("Bldg Score", "building_score", "0", 9),
    ("Econ Score", "economics_score", "0", 9),
    ("Risk Score", "risk_score", "0", 9),
    ("Strategic Fit", "strategic_fit_score", "0", 12),
    ("DD Flags", "due_diligence_flags", None, 40),
    ("Source URL", "source_url", None, 50),
]


def _write_header_row(ws, headers: List[str], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border = _ROW_BORDER


def _autosize_columns(ws, widths: Optional[List[int]] = None) -> None:
    if widths:
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        return
    for col_cells in ws.columns:
        max_len = 0
        letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 55)


def _build_summary_sheet(ws, job: Dict[str, Any], rows: List[Dict[str, Any]], metrics: Dict[str, Any]) -> None:
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60

    # Title banner.
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="K.U.A. — Scan Executive Summary")
    title_cell.font = _TITLE_FONT
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = _ALIGN_LEFT
    ws.row_dimensions[1].height = 26

    def _kv(row: int, key: str, value: Any, number_format: Optional[str] = None) -> None:
        k = ws.cell(row=row, column=1, value=key)
        k.font = Font(bold=True, name="Inter")
        k.alignment = _ALIGN_LEFT
        v = ws.cell(row=row, column=2, value=value)
        v.alignment = _ALIGN_LEFT
        if number_format and isinstance(value, (int, float)):
            v.number_format = number_format

    started = job.get("started_at") or job.get("created_at")
    finished = job.get("finished_at")
    _kv(3, "Job ID", job.get("id"))
    _kv(4, "Job Type", job.get("job_type"))
    _kv(5, "Status", job.get("status"))
    _kv(6, "Search URL", job.get("search_url"))
    _kv(7, "Created", str(job.get("created_at") or ""))
    _kv(8, "Started", str(started or ""))
    _kv(9, "Finished", str(finished or ""))
    _kv(10, "Listings Requested", job.get("listing_limit"), "0")
    _kv(11, "Listings Processed", job.get("listings_done"), "0")

    _kv(13, "Approved candidates", metrics.get("approved_count"), "0")
    _kv(14, "Manual review", metrics.get("manual_review_count"), "0")
    _kv(15, "Rejected", metrics.get("rejected_count"), "0")
    _kv(16, "Failed listings", metrics.get("failed_count"), "0")
    _kv(17, "Total scanned", metrics.get("total_count"), "0")

    _kv(19, "Total estimated EBITDA (approved + manual)", metrics.get("total_ebitda"), '€#,##0')
    _kv(20, "Total investment exposure", metrics.get("total_investment"), '€#,##0')
    _kv(21, "Average margin", metrics.get("avg_margin"), "0.0%")
    _kv(22, "Average EBITDA yield", metrics.get("avg_ebitda_yield"), "0.00%")
    _kv(23, "Average true EBITDA yield (with capex)", metrics.get("avg_true_ebitda_yield"), "0.00%")
    _kv(24, "Average payback (yr, true)", metrics.get("avg_payback_years"), "0.0")

    # Top opportunities mini-table.
    ws.cell(row=26, column=1, value="Top opportunities (by score)").font = Font(bold=True, name="Inter", size=12)
    headers = ["#", "Score", "Deal Status", "Address", "EBITDA", "True Yield", "Source"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=27, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _ALIGN_CENTER
    for idx, r in enumerate(metrics.get("top_opportunities") or [], start=1):
        row = 27 + idx
        ws.cell(row=row, column=1, value=r.get("listing_index"))
        s_cell = ws.cell(row=row, column=2, value=r.get("score"))
        s_cell.number_format = "0"
        ws.cell(row=row, column=3, value=r.get("deal_status"))
        ws.cell(row=row, column=4, value=r.get("address") or "—")
        e_cell = ws.cell(row=row, column=5, value=r.get("ebitda"))
        e_cell.number_format = '€#,##0'
        y_cell = ws.cell(row=row, column=6, value=r.get("true_ebitda_yield"))
        y_cell.number_format = "0.00%"
        ws.cell(row=row, column=7, value=r.get("source_url") or "")

    for col in ("A", "B", "C", "D", "E", "F", "G"):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, 18)
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["G"].width = 50


def _build_pipeline_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    ws.title = "Property Pipeline"
    ws.freeze_panes = "C2"

    headers = [c[0] for c in _PIPELINE_COLUMNS]
    accessors = [c[1] for c in _PIPELINE_COLUMNS]
    number_formats = [c[2] for c in _PIPELINE_COLUMNS]
    widths = [c[3] for c in _PIPELINE_COLUMNS]

    _write_header_row(ws, headers)
    for row_idx, record in enumerate(rows, start=2):
        for col_idx, (accessor, fmt) in enumerate(zip(accessors, number_formats), start=1):
            value = record.get(accessor)
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else None)
            if fmt and isinstance(value, (int, float)):
                cell.number_format = fmt
            cell.alignment = _ALIGN_LEFT if not fmt else _ALIGN_CENTER
            cell.border = _ROW_BORDER
        # Color-code the deal-status column.
        deal_cell = ws.cell(row=row_idx, column=3)
        fill = _DEAL_STATUS_FILLS.get(record.get("deal_status") or "")
        if fill is not None:
            deal_cell.fill = fill

    _autosize_columns(ws, widths)

    # Auto-filter on the header row.
    if rows:
        end_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{end_col}{len(rows) + 1}"


def _build_financial_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    ws.title = "Financial Model"
    ws.freeze_panes = "C2"
    headers = [
        "#",
        "Address",
        "GBA m²",
        "NRA m²",
        "NRA Eff.",
        "Asking Price",
        "CapEx",
        "Total Investment",
        "Annual Revenue",
        "Annual OpEx",
        "EBITDA",
        "Margin",
        "EBITDA Yield",
        "True EBITDA Yield",
        "Payback (yr)",
        "True Payback (yr)",
    ]
    accessors = [
        "listing_index",
        "address",
        "gba_m2",
        "nra_m2",
        "nra_efficiency",
        "asking_price",
        "conversion_capex",
        "total_investment",
        "annual_revenue",
        "annual_opex",
        "ebitda",
        "margin",
        "ebitda_yield",
        "true_ebitda_yield",
        "payback_years",
        "true_payback_years",
    ]
    number_formats = [
        "0",
        None,
        "#,##0",
        "#,##0",
        "0.0%",
        '€#,##0',
        '€#,##0',
        '€#,##0',
        '€#,##0',
        '€#,##0',
        '€#,##0',
        "0.0%",
        "0.00%",
        "0.00%",
        "0.0",
        "0.0",
    ]
    _write_header_row(ws, headers)
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (accessor, fmt) in enumerate(zip(accessors, number_formats), start=1):
            value = r.get(accessor)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fmt and isinstance(value, (int, float)):
                cell.number_format = fmt
            cell.border = _ROW_BORDER
    _autosize_columns(ws, [6, 36, 10, 10, 9, 16, 14, 16, 16, 14, 14, 9, 13, 14, 12, 14])


def _build_memo_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    ws.title = "AI Memos"
    ws.freeze_panes = "A2"
    headers = ["#", "Deal Status", "Score", "Address", "IC Memo"]
    _write_header_row(ws, headers)
    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r.get("listing_index"))
        s = ws.cell(row=row_idx, column=2, value=r.get("deal_status"))
        fill = _DEAL_STATUS_FILLS.get(r.get("deal_status") or "")
        if fill is not None:
            s.fill = fill
        ws.cell(row=row_idx, column=3, value=r.get("score"))
        ws.cell(row=row_idx, column=4, value=r.get("address"))
        memo_cell = ws.cell(row=row_idx, column=5, value=_clean_text(r.get("ic_memo"), limit=30_000))
        memo_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 220
    _autosize_columns(ws, [6, 18, 8, 36, 120])


def build_excel_bytes(
    job: Dict[str, Any],
    listing_rows: List[Dict[str, Any]],
) -> bytes:
    """Generate the full 4-sheet workbook for a job and return raw xlsx bytes.

    Sheets:
      1. Executive Summary
      2. Property Pipeline
      3. Financial Model
      4. AI Memos
    """
    flat = [flatten_result(r) for r in (listing_rows or []) if isinstance(r, dict)]
    metrics = _aggregate_metrics(flat)

    wb = Workbook()
    summary_ws = wb.active
    _build_summary_sheet(summary_ws, job or {}, flat, metrics)
    _build_pipeline_sheet(wb.create_sheet(), flat)
    _build_financial_sheet(wb.create_sheet(), flat)
    _build_memo_sheet(wb.create_sheet(), flat)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def build_csv_bytes(listing_rows: List[Dict[str, Any]]) -> bytes:
    flat = [flatten_result(r) for r in (listing_rows or []) if isinstance(r, dict)]
    if not flat:
        flat = [{c[1]: None for c in _PIPELINE_COLUMNS}]
    fieldnames = [c[1] for c in _PIPELINE_COLUMNS]
    # Include any extra fields we keep on the flattened record.
    extras = [
        "ic_memo",
        "description",
        "latitude",
        "longitude",
        "created_at",
        "updated_at",
        "error",
        "property_id",
    ]
    fieldnames = fieldnames + [f for f in extras if f not in fieldnames]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore", quoting=csv.QUOTE_MINIMAL)
    writer.writeheader()
    for row in flat:
        writer.writerow({k: ("" if row.get(k) is None else row.get(k)) for k in fieldnames})
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------
def build_json_bytes(
    job: Dict[str, Any],
    listing_rows: List[Dict[str, Any]],
    *,
    logs: Optional[List[Dict[str, Any]]] = None,
    errors: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    flat = [flatten_result(r) for r in (listing_rows or []) if isinstance(r, dict)]
    metrics = _aggregate_metrics(flat)
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "job": job or {},
        "metrics": metrics,
        "listings": flat,
        "raw_listings": listing_rows or [],
        "logs": logs or [],
        "errors": errors or [],
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


# ---------------------------------------------------------------------------
# IC Memo (concatenated markdown)
# ---------------------------------------------------------------------------
def build_memo_bytes(
    job: Dict[str, Any],
    listing_rows: List[Dict[str, Any]],
) -> bytes:
    flat = [flatten_result(r) for r in (listing_rows or []) if isinstance(r, dict)]
    memos: List[str] = []
    job_id = (job or {}).get("id") or ""

    memos.append(f"# K.U.A. Investment Committee Memos")
    memos.append("")
    memos.append(f"**Job ID:** `{job_id}`  ")
    memos.append(f"**Generated:** {datetime.now(timezone.utc).isoformat()}  ")
    memos.append(f"**Total listings:** {len(flat)}  ")
    memos.append("")
    memos.append("---")
    memos.append("")

    for idx, r in enumerate(flat, start=1):
        memo = r.get("ic_memo")
        addr = r.get("address") or r.get("source_url") or f"Listing #{idx}"
        memos.append(f"## {idx}. {addr}")
        memos.append("")
        memos.append(f"- Deal status: **{r.get('deal_status') or '—'}**")
        memos.append(f"- Score: **{r.get('score') if r.get('score') is not None else '—'}/100**")
        memos.append(f"- Verdict: **{r.get('verdict') or '—'}**")
        memos.append(f"- Source: {r.get('source_url') or '—'}")
        memos.append("")
        if memo:
            memos.append(memo)
        else:
            memos.append("_No memo generated for this listing._")
        memos.append("")
        memos.append("---")
        memos.append("")

    return "\n".join(memos).encode("utf-8")


# ---------------------------------------------------------------------------
# ZIP package (everything in one archive)
# ---------------------------------------------------------------------------
def build_zip_bytes(
    job: Dict[str, Any],
    listing_rows: List[Dict[str, Any]],
    *,
    logs: Optional[List[Dict[str, Any]]] = None,
    errors: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """Bundle every export format plus raw logs/errors into one .zip."""
    buf = io.BytesIO()
    job_id_seg = _safe_filename_segment((job or {}).get("id"))[:12]
    base = f"kua-scan-{job_id_seg}"

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        try:
            zf.writestr(f"{base}/workbook.xlsx", build_excel_bytes(job, listing_rows))
        except Exception as exc:
            zf.writestr(f"{base}/workbook.xlsx.error.txt", f"Failed to render Excel: {exc}")
        try:
            zf.writestr(f"{base}/pipeline.csv", build_csv_bytes(listing_rows))
        except Exception as exc:
            zf.writestr(f"{base}/pipeline.csv.error.txt", f"Failed to render CSV: {exc}")
        try:
            zf.writestr(
                f"{base}/raw.json",
                build_json_bytes(job, listing_rows, logs=logs, errors=errors),
            )
        except Exception as exc:
            zf.writestr(f"{base}/raw.json.error.txt", f"Failed to render JSON: {exc}")
        try:
            zf.writestr(f"{base}/memos.md", build_memo_bytes(job, listing_rows))
        except Exception as exc:
            zf.writestr(f"{base}/memos.md.error.txt", f"Failed to render memos: {exc}")
        try:
            zf.writestr(f"{base}/job.json", json.dumps(job or {}, indent=2, default=str))
        except Exception:
            pass
        if logs:
            try:
                zf.writestr(f"{base}/logs.json", json.dumps(logs, indent=2, default=str))
            except Exception:
                pass
        if errors:
            try:
                zf.writestr(f"{base}/errors.json", json.dumps(errors, indent=2, default=str))
            except Exception:
                pass
        readme = (
            "K.U.A. SCAN EXPORT PACKAGE\n"
            "==========================\n\n"
            f"Job ID: {(job or {}).get('id')}\n"
            f"Generated: {datetime.now(timezone.utc).isoformat()}\n"
            f"Listings: {len(listing_rows or [])}\n\n"
            "Contents:\n"
            "  workbook.xlsx   Full 4-sheet underwriting workbook\n"
            "  pipeline.csv    Flat property pipeline as CSV\n"
            "  raw.json        Structured machine-readable export\n"
            "  memos.md        Concatenated IC memos (markdown)\n"
            "  job.json        Job metadata\n"
            "  logs.json       Worker logs (if available)\n"
            "  errors.json     Job-level errors (if any)\n"
        )
        zf.writestr(f"{base}/README.txt", readme)

    return buf.getvalue()


# ---------------------------------------------------------------------------
# Single dispatch entrypoint used by the storage layer / workers.
# ---------------------------------------------------------------------------
def build_export(
    fmt: ExportFormat,
    job: Dict[str, Any],
    listing_rows: List[Dict[str, Any]],
    *,
    logs: Optional[List[Dict[str, Any]]] = None,
    errors: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    fmt = (fmt or "").lower()
    if fmt == "excel":
        return build_excel_bytes(job, listing_rows)
    if fmt == "csv":
        return build_csv_bytes(listing_rows)
    if fmt == "json":
        return build_json_bytes(job, listing_rows, logs=logs, errors=errors)
    if fmt == "memo":
        return build_memo_bytes(job, listing_rows)
    if fmt == "zip":
        return build_zip_bytes(job, listing_rows, logs=logs, errors=errors)
    raise ValueError(f"Unknown export format: {fmt!r}")


# Silence unused-import linters for the Table/TableStyleInfo we keep available
# for future enhancements (named ranges, sparklines).
_ = (Table, TableStyleInfo)
